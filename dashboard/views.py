from django.views.generic import CreateView, DetailView, DeleteView, ListView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.urls import reverse_lazy
from django.http import Http404, HttpResponse
from django.db.models import Sum
from django.shortcuts import get_object_or_404

from io import BytesIO
from openpyxl import Workbook

from .models import Category, Expense


class CategoryListView(LoginRequiredMixin, ListView):
    template_name = 'dashboard/index.html'
    context_object_name = 'categories'

    def get_queryset(self):
        author = self.request.user
        queryset = author.categories.all().annotate(
            total_sum=Sum('expenses__total'))
        return queryset


@login_required
def get_report(request):
    try:

        wb = Workbook()
        ws = wb.active
        ws.title = "Expense Report"

        author = request.user
        queryset = Expense.objects.filter(category__author=author)

        for idx, expense in enumerate(queryset, start=1):
            category = get_object_or_404(Category, id=expense.category_id)

            ws.cell(row=idx, column=1, value=expense.title)
            ws.cell(row=idx, column=2, value=category.name)
            ws.cell(row=idx, column=3, value=expense.total)
            ws.cell(row=idx, column=4, value=expense.date)
            ws.cell(row=idx, column=5, value=expense.comment)

        output = BytesIO()
        
        wb.save(output)

        response = HttpResponse(
            output.getvalue(),
            content_type=(f'application/vnd.openxmlformats-officedocument.'
                          f'spreadsheetml.sheet')
        )
        response['Content-Disposition'] = (f'attachment; '
                                           f'filename="expense_report.xlsx"')
        return response

    except Expense.DoesNotExist:
        return HttpResponse("Список расходов пуст.")

    except Category.DoesNotExist:
        return HttpResponse("Список категориев пуст")

    except Exception as e:
        return HttpResponse("Ошибка: " + str(e))


class CategoryDetailView(LoginRequiredMixin, DetailView):
    model = Category
    template_name = 'dashboard/category.html'


class ExpenseDeleteView(LoginRequiredMixin, DeleteView):
    model = Expense
    success_url = reverse_lazy('index')

    def get_object(self, queryset=None):
        expense = super(ExpenseDeleteView, self).get_object(queryset)
        if expense.category.author != self.request.user:
            raise Http404("Вы не можете удалить данный расход")
        return expense


class CategoryDeleteView(LoginRequiredMixin, DeleteView):
    model = Category
    success_url = reverse_lazy('index')

    def get_object(self, queryset=None):
        category = super(CategoryDeleteView, self).get_object(queryset)
        if category.author != self.request.user:
            raise Http404("Вы не можете удалить данную категорию")
        return category


class CategoryCreateView(LoginRequiredMixin, CreateView):
    model = Category
    fields = ('name',)
    template_name = 'dashboard/add_category.html'
    success_url = reverse_lazy('index')

    def form_valid(self, form):
        form.instance.author = self.request.user
        return super().form_valid(form)


class ExpenseCreateView(LoginRequiredMixin, CreateView):
    model = Expense
    fields = '__all__'
    template_name = 'dashboard/add_expense.html'
    success_url = reverse_lazy('index')

    def form_valid(self, form):
        form.instance.author = self.request.user
        return super().form_valid(form)
